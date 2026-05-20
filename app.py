"""Flask HTTP layer for XlumaQR.

This module is intentionally thin: it parses and validates form input,
delegates to the pure-Python helpers in :mod:`qr_generator`, and streams
the resulting PNG / ZIP / PDF bytes back to the client. All real work
lives in ``qr_generator`` so the core remains unit-testable without
spinning up a server.
"""

from __future__ import annotations

import base64
import io
import json
import re
import warnings
import zipfile
from typing import Any, Iterator

from flask import Flask, Response, jsonify, render_template, request, send_file, stream_with_context
from PIL import Image, UnidentifiedImageError

from qr_generator import (
    MAX_BORDER,
    MAX_BOX_SIZE,
    MAX_DATA_LENGTH,
    MAX_LOGO_BYTES,
    MAX_LOGO_DIMENSION,
    LOGO_HARD_MAX_DIMENSION,
    MAX_PADDING,
    MAX_BIB_BATCH_SIZE,
    compute_range,
    generate_qr,
    generate_qr_eps,
    generate_qr_print_png,
    generate_qr_svg,
    generate_sequence,
    generate_sequence_render_plan,
    generate_sequence_svg,
    generate_bib_batch,
    get_template,
    images_to_pdf,
    images_to_zip,
    iter_batch_vector_with_progress,
    iter_batch_with_progress,
    list_templates,
    render_template_preview,
)

app = Flask(__name__)

# Per-process cache of rendered template preview PNGs. Keyed by template
# id, populated lazily on first ``GET /api/qr/templates/<id>/preview``.
# Under Vercel each warm Lambda instance reuses this dict across
# requests, so subsequent gallery loads in the same instance render the
# bytes once and ship the cached payload thereafter. A cold start
# starts an empty dict and re-renders on first request, which is fine.
_PREVIEW_CACHE: dict[str, bytes] = {}

# ``prefix`` is concatenated unmodified into ZIP entry names. Restrict it
# to a conservative set so a hostile caller cannot sneak path separators,
# NULs, or leading dots into the archive. The class admits letters,
# digits, ``_``, ``-``, ``.``, and space so common real-world prefixes
# (``inv.001-``, ``tickets.``, ``2026 batch ``) are accepted, but any
# leading ``.`` is rejected to keep hidden-file names and traversal
# patterns (``../``) out.
_PREFIX_RE = re.compile(r"^(?![.])[A-Za-z0-9_. -]*$")


def _parse_int(
    value: Any,
    field: str,
    *,
    default: int | None = None,
    min_value: int | None = None,
    max_value: int | None = None,
) -> int | None:
    """Coerce a form value to int, returning ``default`` when blank/missing.

    Raises ``ValueError`` with a human-friendly message when the value is
    present but not a valid integer, or falls outside ``[min_value,
    max_value]`` when those bounds are supplied.
    """
    if value is None:
        return default
    if isinstance(value, str):
        if value.strip() == "":
            return default
        try:
            parsed = int(value.strip())
        except ValueError as exc:
            raise ValueError(f"{field} must be an integer") from exc
    else:
        try:
            parsed = int(value)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"{field} must be an integer") from exc

    if min_value is not None and parsed < min_value:
        raise ValueError(f"{field} must be >= {min_value}")
    if max_value is not None and parsed > max_value:
        raise ValueError(f"{field} must be <= {max_value}")
    return parsed


@app.route("/", methods=["GET"])
def index() -> str:
    """Render the single-page UI with Single QR + Sequential Batch tabs."""
    return render_template("index.html")


def _resolve_template_id_from_request() -> str | None:
    """Return the template id to use for this request, or ``None``.

    Reads the optional ``template_id`` form field. Whitespace-only values
    and the literal ``"default"`` are normalised to ``None`` so the
    legacy plain-rendering fast path in :func:`generate_qr` is hit
    byte-for-byte. Any other value is validated against the template
    registry; an unknown id raises :class:`ValueError` with a message
    starting ``"unknown template_id"`` so the route's existing
    ``ValueError -> 400`` handlers can surface it cleanly.
    """
    raw = request.form.get("template_id")
    if raw is None:
        return None
    value = raw.strip()
    if value == "" or value == "default":
        return None
    try:
        get_template(value)
    except ValueError as exc:
        raise ValueError(f"unknown template_id: {value}") from exc
    return value


def _load_logo_from_request() -> Image.Image | None:
    """Load and validate an optional ``logo`` upload from the request.

    Returns ``None`` if no ``logo`` file field is supplied or the field
    is empty (no filename / zero bytes). Otherwise returns the decoded
    :class:`PIL.Image.Image` ready to pass straight into
    :func:`generate_qr`.

    Validation steps, in order, all surfaced as :class:`ValueError` so
    the route's existing ``ValueError -> 400`` handler returns a clean
    JSON 400 (never a 500):

    1. The upload is read with a hard byte cap of
       :data:`qr_generator.MAX_LOGO_BYTES`. Anything larger raises
       ``"logo too large"``. This is a *hard reject* and is not
       relaxed by the auto-resize behaviour below.
    2. The bytes are sniffed via :func:`PIL.Image.open` followed by
       :meth:`PIL.Image.Image.verify`, which validates the magic bytes
       match a real image. ``verify()`` consumes the image object, so
       the bytes are reopened afterwards for further inspection.
    3. The lazy :func:`PIL.Image.open` exposes ``image.size`` *without*
       allocating the full bitmap. Either dimension exceeding
       :data:`qr_generator.LOGO_HARD_MAX_DIMENSION` (the absolute
       ceiling) raises ``"logo dimensions too large"`` *before*
       :meth:`Image.load` runs, so a small-on-the-wire / huge-when-decoded
       PNG (a "decompression bomb") can never exhaust memory on the
       way to the dimension check. Dimensions above
       :data:`qr_generator.MAX_LOGO_DIMENSION` but within the hard
       ceiling are *not* rejected: they are flagged for an auto-resize
       step that runs after :meth:`Image.load` so the user can drop a
       phone-camera screenshot into the form without thinking about
       pixel sizes. PIL's built-in :class:`PIL.Image.DecompressionBombWarning`
       only fires above ~89 MP and is a warning, not an exception, so
       we cannot rely on it.
    4. Only PNG and JPEG images are accepted. Anything else (SVG, BMP,
       GIF, WEBP, etc.) raises ``"logo must be PNG or JPEG"``. The
       format check runs *before* the auto-resize step because
       :meth:`PIL.Image.Image.thumbnail` clears ``image.format``.
    5. If the lazy size check flagged the upload for auto-resize, the
       decoded image is shrunk in-place via
       ``image.thumbnail((MAX_LOGO_DIMENSION, MAX_LOGO_DIMENSION),
       Image.LANCZOS)`` so the working bitmap returned to the caller
       always fits inside the resize target while preserving aspect
       ratio.

    The byte cap (step 1) and the PIL :class:`PIL.Image.DecompressionBombError`
    catch (step 3) remain *hard rejects*; only the soft dimension cap
    (``MAX_LOGO_DIMENSION``) is relaxed by the auto-resize behaviour.

    PIL's :class:`PIL.UnidentifiedImageError` (raised when a non-image
    file is uploaded with a fake mime type, e.g. a text file) and the
    grab-bag of decode-time errors PIL raises on malformed input
    (``OSError``, ``SyntaxError``, ``ValueError``) are caught and turned
    into ``ValueError("logo could not be decoded")``. ``MemoryError``
    is deliberately *not* caught here so an out-of-memory decode
    (which can no longer be triggered through this validator now that
    the dimension check runs first, but could conceivably surface from
    a pathological PIL bug) propagates as a real failure rather than
    being masked as a generic 400.

    :class:`PIL.Image.DecompressionBombError` is the *exception* PIL
    raises (above ~178 MP, twice the warning threshold) from inside
    :func:`Image.open` itself, *before* our own ``image.size`` check
    can run. ``DecompressionBombError`` inherits directly from
    :class:`Exception`, not :class:`OSError`/:class:`SyntaxError`/
    :class:`ValueError`, so it is caught explicitly here and surfaced
    as the ``"logo dimensions too large"`` 400. Such an upload is
    necessarily larger than our own ``LOGO_HARD_MAX_DIMENSION`` ceiling
    (4096 per side), so the dimension-cap message references the hard
    ceiling rather than the auto-resize target.
    """
    upload = request.files.get("logo")
    if upload is None or not getattr(upload, "filename", ""):
        return None

    raw = upload.read()
    if not raw:
        return None
    if len(raw) > MAX_LOGO_BYTES:
        raise ValueError(f"logo too large (max {MAX_LOGO_BYTES} bytes)")

    # First pass: ``verify()`` confirms the magic bytes match a real
    # image without decoding the full bitmap. It also consumes the
    # image object, so we re-open the bytes for further inspection.
    # We silence PIL's ``DecompressionBombWarning`` here: it fires on
    # large headers (above ~89 MP) but is informational, not an error.
    # Our own ``LOGO_HARD_MAX_DIMENSION`` ceiling below is the
    # authoritative hard reject; uploads above ``MAX_LOGO_DIMENSION``
    # but within the hard ceiling are auto-resized, not rejected.
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", Image.DecompressionBombWarning)
            probe = Image.open(io.BytesIO(raw))
            probe.verify()
    except Image.DecompressionBombError as exc:
        # Above ~178 MP (2 * MAX_IMAGE_PIXELS), PIL escalates the bomb
        # warning to an exception inside ``Image.open()`` itself, before
        # our own ``image.size`` check has a chance to run. Map it to
        # the dimension-cap error: such an upload is necessarily larger
        # than our own ``LOGO_HARD_MAX_DIMENSION`` ceiling. Caught
        # explicitly because ``DecompressionBombError`` extends
        # ``Exception`` directly, not the narrower types below.
        raise ValueError(
            f"logo dimensions too large (max {LOGO_HARD_MAX_DIMENSION}x{LOGO_HARD_MAX_DIMENSION})"
        ) from exc
    except UnidentifiedImageError as exc:
        raise ValueError("logo could not be decoded") from exc
    except (OSError, SyntaxError, ValueError) as exc:
        # PIL raises a grab-bag of exceptions on malformed input
        # (SyntaxError on truncated headers, OSError on broken streams,
        # ValueError on other malformed input). Normalise them to a
        # ValueError so the HTTP layer returns 400. ``MemoryError`` and
        # any unexpected ``Exception`` propagate.
        raise ValueError("logo could not be decoded") from exc

    # Second pass: open lazily so ``image.size`` is available without
    # decoding the full bitmap. Reject dimensions above the hard
    # ceiling *before* calling ``load()`` so a decompression-bomb
    # upload (small on the wire, huge on decode) is rejected before
    # allocating the bitmap. Dimensions within the hard ceiling but
    # above the soft auto-resize target are flagged here and shrunk
    # after ``load()`` and the format check.
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", Image.DecompressionBombWarning)
            image = Image.open(io.BytesIO(raw))
    except Image.DecompressionBombError as exc:
        raise ValueError(
            f"logo dimensions too large (max {LOGO_HARD_MAX_DIMENSION}x{LOGO_HARD_MAX_DIMENSION})"
        ) from exc
    except UnidentifiedImageError as exc:
        raise ValueError("logo could not be decoded") from exc
    except (OSError, SyntaxError, ValueError) as exc:
        raise ValueError("logo could not be decoded") from exc

    width, height = image.size
    if width > LOGO_HARD_MAX_DIMENSION or height > LOGO_HARD_MAX_DIMENSION:
        raise ValueError(
            f"logo dimensions too large (max {LOGO_HARD_MAX_DIMENSION}x{LOGO_HARD_MAX_DIMENSION})"
        )
    needs_resize = width > MAX_LOGO_DIMENSION or height > MAX_LOGO_DIMENSION

    # Now that the dimension cap has bounded the worst-case bitmap to
    # ``LOGO_HARD_MAX_DIMENSION ** 2 * 4`` bytes (~64 MB at the 4096
    # ceiling with an RGBA decode), decode the bitmap. The auto-resize
    # below runs *after* ``load()`` so this bound holds for every
    # accepted upload.
    try:
        image.load()
    except UnidentifiedImageError as exc:
        raise ValueError("logo could not be decoded") from exc
    except (OSError, SyntaxError, ValueError) as exc:
        raise ValueError("logo could not be decoded") from exc

    fmt = (image.format or "").upper()
    if fmt not in {"PNG", "JPEG"}:
        raise ValueError("logo must be PNG or JPEG")

    # Auto-resize last: ``Image.thumbnail`` clears ``image.format``, so
    # this step has to follow the format check above. Anything that
    # passed the hard-ceiling check but is above the soft auto-resize
    # target is shrunk in-place to fit inside ``MAX_LOGO_DIMENSION`` on
    # both axes while preserving aspect ratio.
    if needs_resize:
        image.thumbnail(
            (MAX_LOGO_DIMENSION, MAX_LOGO_DIMENSION), Image.LANCZOS
        )

    return image


@app.route("/api/qr/single", methods=["POST"])
def api_single() -> Response:
    """Render a single QR code and return it as a PNG or SVG response.

    ``output_format`` (optional, default ``png``) selects the rendered
    artefact:

    * ``png`` (default): the legacy raster path. The on-screen preview
      uses this so the styled PIL renderer's pixel output is what the
      user sees.
    * ``svg``: the vector path (FEAT-002). Returns an
      ``image/svg+xml`` body whose QR pattern is fully vector and so
      stays sharp at any zoom level. The HD download button on the
      Single QR and Batch live previews uses this format. Embedded
      logos are encoded as base64 PNG data URIs (small region, raster
      trade-off documented in :func:`qr_generator.generate_qr_svg`).
    """
    data = request.form.get("data", "").strip()
    if not data:
        return jsonify({"error": "data is required"}), 400
    if len(data) > MAX_DATA_LENGTH:
        return jsonify({"error": f"data must be <= {MAX_DATA_LENGTH} characters"}), 400

    label = request.form.get("label")
    if label is not None and label == "":
        label = None

    output_format = (request.form.get("output_format") or "png").strip().lower()
    if output_format not in {"png", "svg", "eps", "print_png"}:
        return jsonify({"error": "output_format must be 'png', 'svg', 'eps', or 'print_png'"}), 400

    try:
        box_size = _parse_int(
            request.form.get("box_size"),
            "box_size",
            default=10,
            min_value=1,
            max_value=MAX_BOX_SIZE,
        )
        border = _parse_int(
            request.form.get("border"),
            "border",
            default=4,
            min_value=0,
            max_value=MAX_BORDER,
        )
        template_id = _resolve_template_id_from_request()
        logo = _load_logo_from_request()
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    assert box_size is not None and border is not None  # defaults guarantee non-None

    if output_format == "eps":
        try:
            eps_bytes = generate_qr_eps(
                data,
                label=label,
                box_size=box_size,
                border=border,
                template_id=template_id,
                logo=logo,
            )
        except ValueError as exc:
            return jsonify({"error": f"data could not be encoded: {exc}"}), 400
        return send_file(
            io.BytesIO(eps_bytes),
            mimetype="application/postscript",
            as_attachment=True,
            download_name="qr.eps",
        )

    if output_format == "print_png":
        try:
            png_bytes = generate_qr_print_png(
                data,
                label=label,
                box_size=40,
                border=border,
                template_id=template_id,
                logo=logo,
                dpi=300,
            )
        except ValueError as exc:
            return jsonify({"error": f"data could not be encoded: {exc}"}), 400
        return send_file(
            io.BytesIO(png_bytes),
            mimetype="image/png",
            as_attachment=True,
            download_name="qr_300dpi.png",
        )

    if output_format == "svg":
        try:
            svg = generate_qr_svg(
                data,
                label=label,
                box_size=box_size,
                border=border,
                template_id=template_id,
                logo=logo,
            )
        except ValueError as exc:
            return jsonify({"error": f"data could not be encoded: {exc}"}), 400
        return send_file(
            io.BytesIO(svg.encode("utf-8")),
            mimetype="image/svg+xml",
            as_attachment=False,
            download_name="qr.svg",
        )

    try:
        image = generate_qr(
            data,
            label=label,
            box_size=box_size,
            border=border,
            template_id=template_id,
            logo=logo,
        )
    except ValueError as exc:
        # ``qrcode`` raises ValueError when the encoded payload exceeds
        # version 40 capacity (e.g. multi-byte UTF-8 input under the cap
        # but over the byte ceiling). Surface as a 400, not a 500.
        return jsonify({"error": f"data could not be encoded: {exc}"}), 400
    buf = io.BytesIO()
    image.save(buf, format="PNG")
    buf.seek(0)
    return send_file(
        buf,
        mimetype="image/png",
        as_attachment=False,
        download_name="qr.png",
    )


@app.route("/api/qr/batch", methods=["POST"])
def api_batch() -> Response:
    """Render a sequential batch and return a ZIP or PDF response.

    The ``format`` field selects the artefact:

    * ``zip`` (legacy): one PNG per code in a ZIP archive. Preserved
      for back-compat; the rendered entries are PIL images.
    * ``zip_svg`` (FEAT-002): one SVG per code in a ZIP archive. Each
      entry's QR pattern is fully vector and so stays sharp at any
      zoom level. Mimetype is ``application/zip``; entries end in
      ``.svg``.
    * ``pdf`` (FEAT-002): a vector PDF where each QR module is drawn
      as a reportlab primitive (no embedded raster image XObjects
      unless a logo is supplied). The grid layout is unchanged.
    """
    parsed, err = _parse_batch_form()
    if err is not None:
        return jsonify({"error": err}), 400
    assert parsed is not None

    fmt = parsed["fmt"]
    first_n = parsed["first_n"]
    last_n = parsed["last_n"]

    try:
        if fmt == "pdf":
            plans = generate_sequence_render_plan(
                start=parsed["start"],
                count=parsed["count"],
                end=parsed["end"],
                data_template=parsed["data_template"],
                label_template=parsed["label_template"],
                padding=parsed["padding"],
                prefix=parsed["prefix"],
                box_size=parsed["box_size"],
                border=parsed["border"],
                template_id=parsed["template_id"],
                logo=parsed["logo"],
            )
            payload: bytes | None = None
            for token in iter_batch_vector_with_progress(plans, "pdf"):
                if token[0] == "result":
                    payload = token[1]
            assert payload is not None
            mimetype = "application/pdf"
            filename = f"qr_batch_{first_n}_{last_n}.pdf"
        elif fmt == "pdf_single":
            plans = generate_sequence_render_plan(
                start=parsed["start"],
                count=parsed["count"],
                end=parsed["end"],
                data_template=parsed["data_template"],
                label_template=parsed["label_template"],
                padding=parsed["padding"],
                prefix=parsed["prefix"],
                box_size=parsed["box_size"],
                border=parsed["border"],
                template_id=parsed["template_id"],
                logo=parsed["logo"],
            )
            from qr_generator import generate_pdf_single_page
            payload = generate_pdf_single_page(plans)
            mimetype = "application/pdf"
            filename = f"qr_batch_{first_n}_{last_n}.pdf"
        elif fmt == "zip_svg":
            svg_items = generate_sequence_svg(
                start=parsed["start"],
                count=parsed["count"],
                end=parsed["end"],
                data_template=parsed["data_template"],
                label_template=parsed["label_template"],
                padding=parsed["padding"],
                prefix=parsed["prefix"],
                box_size=parsed["box_size"],
                border=parsed["border"],
                template_id=parsed["template_id"],
                logo=parsed["logo"],
            )
            payload2: bytes | None = None
            for token in iter_batch_vector_with_progress(svg_items, "zip_svg"):
                if token[0] == "result":
                    payload2 = token[1]
            assert payload2 is not None
            payload = payload2
            mimetype = "application/zip"
            filename = f"qr_batch_{first_n}_{last_n}.zip"
        elif fmt == "zip_eps":
            # Generate EPS files for each QR in the batch
            numbers = compute_range(
                parsed["start"],
                count=parsed["count"],
                end=parsed["end"],
                padding=parsed["padding"],
            )
            eps_buffer = io.BytesIO()
            with zipfile.ZipFile(eps_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
                data_template = parsed["data_template"]
                label_template = parsed["label_template"]
                prefix = parsed["prefix"]
                for n in numbers:
                    qr_data = data_template.replace("{n}", n)
                    qr_label = label_template.replace("{n}", n) if label_template is not None else None
                    eps_bytes = generate_qr_eps(
                        qr_data,
                        label=qr_label,
                        box_size=parsed["box_size"],
                        border=parsed["border"],
                        template_id=parsed["template_id"],
                        logo=parsed["logo"],
                    )
                    zf.writestr(f"{prefix}{n}.eps", eps_bytes)
            payload = eps_buffer.getvalue()
            mimetype = "application/zip"
            filename = f"qr_batch_{first_n}_{last_n}_eps.zip"
        else:
            # PNG batch: render at high resolution (box_size=40) for
            # print quality. The user's box_size setting is overridden
            # to ensure the output doesn't pixelate when printed.
            items = generate_sequence(
                start=parsed["start"],
                count=parsed["count"],
                end=parsed["end"],
                data_template=parsed["data_template"],
                label_template=parsed["label_template"],
                padding=parsed["padding"],
                prefix=parsed["prefix"],
                box_size=40,
                border=parsed["border"],
                template_id=parsed["template_id"],
                logo=parsed["logo"],
            )
            payload = images_to_zip(items)
            mimetype = "application/zip"
            filename = f"qr_batch_{first_n}_{last_n}.zip"
    except ValueError as exc:
        # ``qrcode`` raises ValueError when a substituted template
        # overflows QR version 40 capacity. Surface as a 400.
        return jsonify({"error": f"batch could not be encoded: {exc}"}), 400

    return send_file(
        io.BytesIO(payload),
        mimetype=mimetype,
        as_attachment=True,
        download_name=filename,
    )


def _parse_batch_form() -> tuple[dict[str, Any] | None, str | None]:
    """Validate the batch form and return either parsed values or an error.

    Returns ``(parsed, None)`` on success and ``(None, error_message)``
    on failure. The error message is suitable for a 400 JSON body and
    matches the messages produced by the long-standing
    :func:`api_batch` route, so the synchronous and streaming routes
    return identical 400 contracts for the same bad input.
    """
    try:
        start = _parse_int(request.form.get("start"), "start")
        if start is None:
            return None, "start is required"
        count = _parse_int(request.form.get("count"), "count")
        end = _parse_int(request.form.get("end"), "end")
        padding = _parse_int(
            request.form.get("padding"),
            "padding",
            default=0,
            min_value=0,
            max_value=MAX_PADDING,
        )
        box_size = _parse_int(
            request.form.get("box_size"),
            "box_size",
            default=10,
            min_value=1,
            max_value=MAX_BOX_SIZE,
        )
        border = _parse_int(
            request.form.get("border"),
            "border",
            default=4,
            min_value=0,
            max_value=MAX_BORDER,
        )
        template_id = _resolve_template_id_from_request()
        logo = _load_logo_from_request()
    except ValueError as exc:
        return None, str(exc)

    assert padding is not None and box_size is not None and border is not None

    if count is None and end is None:
        return None, "exactly one of count or end is required"
    if count is not None and end is not None:
        return None, "provide only one of count or end, not both"

    prefix = request.form.get("prefix", "")
    if not _PREFIX_RE.match(prefix):
        return None, (
            "prefix may only contain letters, digits, "
            "'_', '-', '.', or space, and may not start with '.'"
        )
    data_template = request.form.get("data_template") or "{n}"
    if len(data_template) > MAX_DATA_LENGTH:
        return None, f"data_template must be <= {MAX_DATA_LENGTH} characters"

    raw_label_template = request.form.get("label_template")
    if raw_label_template is None:
        label_template: str | None = "{n}"
    elif raw_label_template == "":
        label_template = None
    else:
        label_template = raw_label_template
    if label_template is not None and len(label_template) > MAX_DATA_LENGTH:
        return None, f"label_template must be <= {MAX_DATA_LENGTH} characters"

    fmt = (request.form.get("format") or "zip").strip().lower()
    if fmt not in {"zip", "zip_svg", "zip_eps", "pdf", "pdf_single"}:
        return None, "format must be 'zip', 'zip_svg', 'zip_eps', 'pdf', or 'pdf_single'"

    # Validate the range up front so we can return a clean 400 before we
    # start rendering hundreds of QR codes.
    try:
        numbers = compute_range(start, count=count, end=end, padding=padding)
    except ValueError as exc:
        return None, str(exc)

    return (
        {
            "start": start,
            "count": count,
            "end": end,
            "padding": padding,
            "box_size": box_size,
            "border": border,
            "prefix": prefix,
            "data_template": data_template,
            "label_template": label_template,
            "fmt": fmt,
            "first_n": numbers[0],
            "last_n": numbers[-1],
            "total": len(numbers),
            "template_id": template_id,
            "logo": logo,
        },
        None,
    )


@app.route("/api/qr/batch/stream", methods=["POST"])
def api_batch_stream() -> Response:
    """Render a sequential batch and stream NDJSON progress events.

    Validation behaves exactly like :func:`api_batch`: invalid input
    returns a 400 JSON body and never enters streaming mode. Once
    validation passes, the response is ``application/x-ndjson`` and the
    body consists of one JSON object per line:

    * ``{"event": "start", "total": N, "format": "zip|pdf",
       "first": "...", "last": "..."}``
    * ``{"event": "progress", "index": i, "total": N, "name": "<filename>"}``
       once per generated QR (``i`` runs ``0..N-1``).
    * Either ``{"event": "result", "filename": "...",
       "mimetype": "application/zip|application/pdf",
       "data_base64": "..."}`` carrying the packed bytes, OR
       ``{"event": "error", "error": "..."}`` if encoding fails
       mid-stream. The HTTP status remains 200 in either case once
       streaming has begun.
    """
    parsed, err = _parse_batch_form()
    if err is not None:
        return jsonify({"error": err}), 400
    assert parsed is not None

    fmt: str = parsed["fmt"]
    total: int = parsed["total"]
    first_n: str = parsed["first_n"]
    last_n: str = parsed["last_n"]

    if fmt == "pdf" or fmt == "pdf_single":
        mimetype = "application/pdf"
        filename = f"qr_batch_{first_n}_{last_n}.pdf"
    else:
        # Both ``zip`` (PNG entries) and ``zip_svg`` (SVG entries) are
        # delivered as ``application/zip`` archives; the entries inside
        # carry the appropriate per-file extension.
        mimetype = "application/zip"
        filename = f"qr_batch_{first_n}_{last_n}.zip"

    if fmt == "zip_svg":
        items = generate_sequence_svg(
            start=parsed["start"],
            count=parsed["count"],
            end=parsed["end"],
            data_template=parsed["data_template"],
            label_template=parsed["label_template"],
            padding=parsed["padding"],
            prefix=parsed["prefix"],
            box_size=parsed["box_size"],
            border=parsed["border"],
            template_id=parsed["template_id"],
            logo=parsed["logo"],
        )
        packer = lambda src: iter_batch_vector_with_progress(src, "zip_svg")
    elif fmt == "pdf" or fmt == "pdf_single":
        items = generate_sequence_render_plan(
            start=parsed["start"],
            count=parsed["count"],
            end=parsed["end"],
            data_template=parsed["data_template"],
            label_template=parsed["label_template"],
            padding=parsed["padding"],
            prefix=parsed["prefix"],
            box_size=parsed["box_size"],
            border=parsed["border"],
            template_id=parsed["template_id"],
            logo=parsed["logo"],
        )
        packer = lambda src: iter_batch_vector_with_progress(src, fmt)
    elif fmt == "zip_eps":
        # EPS batch: generate EPS files and pack into a ZIP with progress
        numbers = compute_range(
            parsed["start"],
            count=parsed["count"],
            end=parsed["end"],
            padding=parsed["padding"],
        )

        def _eps_packer(src):
            """Generate EPS files and yield progress + result tokens."""
            eps_parts = []
            data_template = parsed["data_template"]
            label_template = parsed["label_template"]
            prefix = parsed["prefix"]
            for idx, n in enumerate(numbers):
                qr_data = data_template.replace("{n}", n)
                qr_label = label_template.replace("{n}", n) if label_template is not None else None
                eps_bytes = generate_qr_eps(
                    qr_data,
                    label=qr_label,
                    box_size=parsed["box_size"],
                    border=parsed["border"],
                    template_id=parsed["template_id"],
                    logo=parsed["logo"],
                )
                fname = f"{prefix}{n}.eps"
                eps_parts.append((fname, eps_bytes))
                yield ("progress", idx, fname)
            # Pack into ZIP
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
                for fname, data in eps_parts:
                    zf.writestr(fname, data)
            yield ("result", buf.getvalue())

        items = None
        packer = _eps_packer
    else:
        # PNG batch: render at high resolution (box_size=40) for print
        items = generate_sequence(
            start=parsed["start"],
            count=parsed["count"],
            end=parsed["end"],
            data_template=parsed["data_template"],
            label_template=parsed["label_template"],
            padding=parsed["padding"],
            prefix=parsed["prefix"],
            box_size=40,
            border=parsed["border"],
            template_id=parsed["template_id"],
            logo=parsed["logo"],
        )
        packer = lambda src: iter_batch_with_progress(src, fmt)

    def _ndjson(obj: dict[str, Any]) -> str:
        # ``ensure_ascii`` keeps non-ASCII filenames safe on the wire as
        # \uXXXX escapes; downstream JSON parsers handle them transparently.
        return json.dumps(obj, ensure_ascii=True) + "\n"

    def generate() -> Iterator[str]:
        yield _ndjson(
            {
                "event": "start",
                "total": total,
                "format": fmt,
                "first": first_n,
                "last": last_n,
            }
        )

        # Drive the streaming packer one item at a time. ``items`` is
        # the lazy iterator returned by :func:`generate_sequence`, and
        # ``iter_batch_with_progress`` is itself a generator that pulls
        # one ``(filename, image)`` pair per turn, writes it into the
        # output container, then yields a ``("progress", ...)`` tuple
        # back. This keeps peak memory at roughly one PIL Image rather
        # than buffering all N images while we wait to start packing.
        try:
            payload: bytes | None = None
            for token in packer(items) if items is not None else packer(None):
                kind = token[0]
                if kind == "progress":
                    _, index, name = token
                    yield _ndjson(
                        {
                            "event": "progress",
                            "index": index,
                            "total": total,
                            "name": name,
                        }
                    )
                elif kind == "result":
                    payload = token[1]
        except ValueError as exc:
            yield _ndjson(
                {"event": "error", "error": f"batch could not be encoded: {exc}"}
            )
            return
        except Exception as exc:  # pragma: no cover - defensive catch-all
            yield _ndjson({"event": "error", "error": str(exc)})
            return

        # ``payload`` is set by the terminal ``("result", bytes)`` token
        # the helper yields exactly once. If it is missing here, the
        # helper contract was violated; surface that as an error event
        # rather than a 500.
        if payload is None:  # pragma: no cover - contract violation
            yield _ndjson(
                {"event": "error", "error": "internal error: empty batch payload"}
            )
            return

        yield _ndjson(
            {
                "event": "result",
                "filename": filename,
                "mimetype": mimetype,
                "data_base64": base64.b64encode(payload).decode("ascii"),
            }
        )

    response = Response(
        stream_with_context(generate()),
        mimetype="application/x-ndjson",
    )
    response.headers["Cache-Control"] = "no-cache"
    response.headers["X-Accel-Buffering"] = "no"
    return response


@app.route("/api/qr/templates", methods=["GET"])
def api_templates_list() -> Response:
    """Return the JSON listing of built-in design templates.

    The response shape is ``{"templates": [<entry>, ...]}`` where every
    entry has ``id``, ``name``, ``category``, and ``spec`` keys (the
    ``spec`` is included verbatim so the UI can render a small swatch
    if it wants to, but it is not required to use it).

    A short ``Cache-Control`` is set so warm browsers skip the round
    trip on subsequent page loads inside the cache window.
    """
    response = jsonify({"templates": list_templates()})
    response.headers["Cache-Control"] = "public, max-age=300"
    return response


@app.route("/api/qr/templates/<template_id>/preview", methods=["GET"])
def api_template_preview(template_id: str) -> Response:
    """Return a small thumbnail PNG for ``template_id``.

    Resolves the id via :func:`qr_generator.get_template`; an unknown id
    returns a JSON 404. On success, the PNG bytes are produced by
    :func:`qr_generator.render_template_preview` and cached in the
    per-process :data:`_PREVIEW_CACHE` so a warm Lambda renders each
    template at most once.
    """
    try:
        get_template(template_id)
    except ValueError:
        return jsonify({"error": "unknown template id"}), 404

    payload = _PREVIEW_CACHE.get(template_id)
    if payload is None:
        payload = render_template_preview(template_id)
        _PREVIEW_CACHE[template_id] = payload

    response = send_file(
        io.BytesIO(payload),
        mimetype="image/png",
        as_attachment=False,
        download_name=f"{template_id}.png",
    )
    response.headers["Cache-Control"] = "public, max-age=3600"
    return response


@app.route("/api/qr/bib-batch", methods=["POST"])
def api_bib_batch() -> Response:
    """Generate QR codes with unique codes mapped to bib numbers.

    Accepts a list of bib numbers (which may contain letters, dashes,
    or any characters), generates a unique ``XLUMA-xxxxxxxx`` code for
    each, encodes that code in the QR, and returns a ZIP containing:

    * One QR image per bib (PNG or SVG depending on ``format``)
    * An Excel file (``bib_mapping.xlsx``) with columns:
      ``qr_code``, ``bib_number``

    The mapping file is what gets imported into the Xluma platform so
    the mobile app can resolve scanned QR codes back to the correct
    bib number — regardless of whether the bib contains letters,
    dashes, or other non-numeric characters.

    Form fields:

    * ``bibs`` (required): comma-separated list of bib numbers, OR
      one bib per line. Whitespace around each bib is trimmed.
    * ``box_size`` (default 10): pixel size of each QR module.
    * ``border`` (default 4): quiet-zone width in modules.
    * ``template_id`` (optional): design template slug.
    * ``logo`` (optional): centre logo file (PNG or JPEG).
    * ``label_bibs`` (default ``true``): if ``true``, the bib number
      is printed as the QR label so the sticker is human-readable.
    * ``format`` (default ``zip``): ``zip`` for PNG images, ``zip_svg``
      for SVG vector images.
    """
    raw_bibs = request.form.get("bibs", "").strip()
    if not raw_bibs:
        return jsonify({"error": "bibs is required"}), 400

    # Parse bibs: support both comma-separated and newline-separated
    if "\n" in raw_bibs:
        bibs = [b.strip() for b in raw_bibs.splitlines() if b.strip()]
    else:
        bibs = [b.strip() for b in raw_bibs.split(",") if b.strip()]

    if not bibs:
        return jsonify({"error": "no valid bib numbers provided"}), 400
    if len(bibs) > MAX_BIB_BATCH_SIZE:
        return jsonify({"error": f"too many bibs (max {MAX_BIB_BATCH_SIZE})"}), 400

    # Check for duplicates
    seen: set[str] = set()
    for bib in bibs:
        if bib in seen:
            return jsonify({"error": f"duplicate bib number: {bib}"}), 400
        seen.add(bib)

    try:
        box_size = _parse_int(
            request.form.get("box_size"),
            "box_size",
            default=10,
            min_value=1,
            max_value=MAX_BOX_SIZE,
        )
        border = _parse_int(
            request.form.get("border"),
            "border",
            default=4,
            min_value=0,
            max_value=MAX_BORDER,
        )
        template_id = _resolve_template_id_from_request()
        logo = _load_logo_from_request()
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    assert box_size is not None and border is not None

    label_bibs = request.form.get("label_bibs", "true").strip().lower() != "false"
    fmt = (request.form.get("format") or "zip").strip().lower()
    if fmt not in {"zip", "zip_svg"}:
        return jsonify({"error": "format must be 'zip' or 'zip_svg'"}), 400

    try:
        items, mapping = generate_bib_batch(
            bibs,
            box_size=40,  # Always high-res for print quality
            border=border,
            template_id=template_id,
            logo=logo,
            label_bibs=label_bibs,
        )
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400

    # Build the ZIP with QR images + Excel mapping
    import openpyxl

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        # Add QR images
        for filename, image in items:
            if fmt == "zip_svg":
                # Find the matching bib for this filename to get the unique code
                idx = next(
                    i for i, (fn, _) in enumerate(items) if fn == filename
                )
                entry = mapping[idx]
                svg_label = entry["bib_number"] if label_bibs else None
                svg = generate_qr_svg(
                    entry["qr_code"],
                    label=svg_label,
                    box_size=box_size,
                    border=border,
                    template_id=template_id,
                    logo=logo,
                )
                svg_filename = filename.replace(".png", ".svg")
                zf.writestr(svg_filename, svg.encode("utf-8"))
            else:
                # PNG: make transparent background
                image = image.convert("RGBA")
                gray = image.convert("L")
                alpha = gray.point(lambda p: 255 - p)
                image.putalpha(alpha)
                png_buf = io.BytesIO()
                image.save(png_buf, format="PNG")
                zf.writestr(filename, png_buf.getvalue())

        # Add Excel mapping file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "QR Mapping"
        ws.append(["QR Code", "Bib Number"])
        for entry in mapping:
            ws.append([entry["qr_code"], entry["bib_number"]])

        # Style the header row
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="FF4D2D", end_color="FF4D2D", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Auto-width columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 4

        xlsx_buf = io.BytesIO()
        wb.save(xlsx_buf)
        zf.writestr("bib_mapping.xlsx", xlsx_buf.getvalue())

    zip_buffer.seek(0)
    first_bib = bibs[0].replace("/", "_").replace("\\", "_")
    last_bib = bibs[-1].replace("/", "_").replace("\\", "_")
    filename = f"qr_bibs_{first_bib}_{last_bib}.zip"

    return send_file(
        zip_buffer,
        mimetype="application/zip",
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False)
